from dataclasses import dataclass, field
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape

from django.db import transaction

from .models import CBT, Choice, Question, UBT, UBTChoice, UBTQuestion


CBT_REQUIRED_COLUMNS = [
    "cbt_title", "cbt_description", "passing_score", "question_order", "question_text",
    "choice_a", "choice_b", "choice_c", "choice_d", "correct_choice", "explanation",
]
UBT_REQUIRED_COLUMNS = [
    "ubt_title", "ubt_description", "passing_score", "question_order", "question_text",
    "choice_a", "choice_b", "choice_c", "choice_d", "correct_choice", "explanation",
]
CHOICE_KEYS = ["a", "b", "c", "d"]
CHOICE_TYPES = {"text", "image", "audio", "video"}
REQUIRED_COLUMNS = CBT_REQUIRED_COLUMNS


@dataclass
class CBTImportResult:
    created_cbts: int = 0
    updated_cbts: int = 0
    created_questions: int = 0
    created_choices: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self):
        return not self.errors


def normalize(value):
    return "" if value is None else str(value).strip()


def parse_int(value, default=None):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_excel_rows(uploaded_file, required_columns, title_column):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], CBTImportResult(errors=["Dependency openpyxl belum terpasang. Jalankan: pip install openpyxl"])

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        return [], CBTImportResult(errors=[f"File Excel tidak bisa dibaca: {exc}"])

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], CBTImportResult(errors=["File Excel kosong."])

    headers = [normalize(value).lower() for value in rows[0]]
    missing = [column for column in required_columns if column not in headers]
    if missing:
        return [], CBTImportResult(errors=["Kolom wajib belum ada: " + ", ".join(missing)])

    index = {column: headers.index(column) for column in required_columns}
    all_index = {column: position for position, column in enumerate(headers)}
    parsed_rows = []
    result = CBTImportResult()

    for excel_row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(normalize(value) for value in row):
            continue

        def value(column):
            position = index[column]
            return row[position] if position < len(row) else ""

        def optional_value(column):
            if column not in all_index:
                return ""
            position = all_index[column]
            return row[position] if position < len(row) else ""

        item_title = normalize(value(title_column))
        item_description = normalize(value(required_columns[1]))
        passing_score = parse_int(value("passing_score"), 70)
        question_order = parse_int(value("question_order"), len(parsed_rows) + 1)
        question_text = normalize(value("question_text"))
        media_type = normalize(optional_value("media_type")).lower() or "none"
        media_url = normalize(optional_value("media_url"))
        correct_choice = normalize(value("correct_choice")).upper()
        explanation = normalize(value("explanation"))
        choices = []

        for key in CHOICE_KEYS:
            choice_type = normalize(optional_value(f"choice_{key}_type")).lower() or "text"
            choice_text = normalize(value(f"choice_{key}"))
            choice_media_url = normalize(optional_value(f"choice_{key}_media_url"))
            choices.append({
                "text": choice_text,
                "answer_type": choice_type,
                "media_url": choice_media_url,
            })

        row_errors = []
        if not item_title:
            row_errors.append(f"{title_column} wajib diisi")
        if passing_score is None or not 0 <= passing_score <= 100:
            row_errors.append("passing_score harus angka 0-100")
        if question_order is None or question_order < 1:
            row_errors.append("question_order harus angka minimal 1")
        if not question_text:
            row_errors.append("question_text wajib diisi")
        if media_type not in {"none", "image", "audio", "video", "youtube"}:
            row_errors.append("media_type harus none, image, audio, video, atau youtube")
        if media_type in {"image", "audio", "video", "youtube"} and not media_url:
            row_errors.append("media_url wajib diisi jika media_type bukan none")

        filled_choices = []
        for index_number, choice in enumerate(choices, start=1):
            if choice["answer_type"] not in CHOICE_TYPES:
                row_errors.append(f"choice_{CHOICE_KEYS[index_number - 1]}_type harus text, image, audio, atau video")
                continue
            if choice["answer_type"] == "text":
                has_choice = bool(choice["text"])
                choice["media_url"] = ""
            else:
                has_choice = bool(choice["media_url"])
                if has_choice is False and choice["text"]:
                    row_errors.append(f"choice_{CHOICE_KEYS[index_number - 1]}_media_url wajib diisi untuk pilihan media")
            if has_choice:
                filled_choices.append(index_number)

        if len(filled_choices) < 2:
            row_errors.append("minimal 2 pilihan jawaban wajib diisi")
        if correct_choice not in {"A", "B", "C", "D"}:
            row_errors.append("correct_choice harus A, B, C, atau D")
        elif ord(correct_choice) - ord("A") + 1 not in filled_choices:
            row_errors.append("pilihan yang ditandai benar tidak boleh kosong")

        if row_errors:
            result.errors.append(f"Baris {excel_row_number}: " + "; ".join(row_errors))
            continue

        parsed_rows.append({
            "title": item_title,
            "description": item_description,
            "passing_score": passing_score,
            "question_order": question_order,
            "question_text": question_text,
            "media_type": media_type,
            "media_url": media_url,
            "choices": choices,
            "correct_choice": correct_choice,
            "explanation": explanation,
        })

    if result.errors:
        return [], result
    if not parsed_rows:
        return [], CBTImportResult(errors=["Tidak ada baris soal yang bisa diimport."])
    return parsed_rows, result


def save_imported_rows(rows, mode, parent_model, question_model, choice_model, relation_name):
    result = CBTImportResult()
    touched_ids = set()
    with transaction.atomic():
        for item in rows:
            parent, created = parent_model.objects.get_or_create(
                title=item["title"],
                defaults={"description": item["description"], "passing_score": item["passing_score"], "is_active": True},
            )
            if created:
                result.created_cbts += 1
            else:
                result.updated_cbts += 1 if parent.id not in touched_ids else 0
                parent.description = item["description"]
                parent.passing_score = item["passing_score"]
                parent.is_active = True
                parent.save(update_fields=["description", "passing_score", "is_active"])

            if mode == "replace" and parent.id not in touched_ids:
                getattr(parent, relation_name).all().delete()
            touched_ids.add(parent.id)

            question_kwargs = {
                relation_name[:-1] if relation_name.endswith("s") else relation_name: parent,
                "text": item["question_text"],
                "explanation": item["explanation"],
                "media_type": item["media_type"],
                "media_url": item["media_url"],
                "order": item["question_order"],
            }
            if parent_model is CBT:
                question_kwargs = {"cbt": parent, **{k: v for k, v in question_kwargs.items() if k not in {"question"}}}
            else:
                question_kwargs = {"ubt": parent, **{k: v for k, v in question_kwargs.items() if k not in {"question"}}}
            question = question_model.objects.create(**question_kwargs)
            result.created_questions += 1

            correct_index = ord(item["correct_choice"]) - ord("A")
            for index_number, choice in enumerate(item["choices"]):
                has_choice = bool(choice["text"]) if choice["answer_type"] == "text" else bool(choice["media_url"])
                if not has_choice:
                    continue
                choice_model.objects.create(
                    question=question,
                    text=choice["text"],
                    answer_type=choice["answer_type"],
                    media_url=choice["media_url"],
                    is_correct=index_number == correct_index,
                )
                result.created_choices += 1
    return result


def import_cbt_from_excel(uploaded_file, mode):
    rows, result = parse_excel_rows(uploaded_file, CBT_REQUIRED_COLUMNS, "cbt_title")
    if result.errors:
        return result
    return save_imported_rows(rows, mode, CBT, Question, Choice, "questions")


def import_ubt_from_excel(uploaded_file, mode):
    rows, result = parse_excel_rows(uploaded_file, UBT_REQUIRED_COLUMNS, "ubt_title")
    if result.errors:
        return result
    return save_imported_rows(rows, mode, UBT, UBTQuestion, UBTChoice, "questions")


def build_template_xlsx(headers, sample_rows, sheet_name):
    rows = [headers, *sample_rows]

    def col_name(index):
        name = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name
        return name

    def cell_xml(row_number, col_number, value):
        reference = f"{col_name(col_number)}{row_number}"
        return f'<c r="{reference}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>'

    sheet_rows = []
    for row_number, row in enumerate(rows, start=1):
        cells = "".join(cell_xml(row_number, col_number, value) for col_number, value in enumerate(row, start=1))
        sheet_rows.append(f'<row r="{row_number}">{cells}</row>')

    last_col = col_name(len(headers))
    sheet_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <dimension ref="A1:{last_col}{len(rows)}"/>
  <sheetData>{''.join(sheet_rows)}</sheetData>
</worksheet>'''

    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>''')
        archive.writestr("_rels/.rels", '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>''')
        archive.writestr("xl/workbook.xml", f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="{escape(sheet_name)}" sheetId="1" r:id="rId1"/></sheets></workbook>''')
        archive.writestr("xl/_rels/workbook.xml.rels", '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>''')
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)

    output.seek(0)
    return output.getvalue()


def choice_media_headers(prefix):
    return [
        f"{prefix}",
        f"{prefix}_type",
        f"{prefix}_media_url",
    ]


def build_import_headers(required_columns):
    return (
        required_columns[:5]
        + ["media_type", "media_url"]
        + choice_media_headers("choice_a")
        + choice_media_headers("choice_b")
        + choice_media_headers("choice_c")
        + choice_media_headers("choice_d")
        + ["correct_choice", "explanation"]
    )


def build_cbt_import_template_xlsx():
    headers = build_import_headers(CBT_REQUIRED_COLUMNS)
    samples = [
        ["CBT Kosakata Korea", "Latihan kosakata dasar", "70", "1", "Apa arti sekolah?", "none", "", "Sekolah", "text", "", "Rumah", "text", "", "Pasar", "text", "", "Kantor", "text", "", "A", "hakgyo berarti sekolah."],
        ["CBT Kosakata Korea", "Latihan kosakata dasar", "70", "2", "Dengarkan audio, mana arti yang tepat?", "audio", "https://example.com/audio/mul.mp3", "Air", "text", "", "", "image", "https://example.com/images/book.png", "Makanan", "text", "", "Teman", "text", "", "A", "mul berarti air."],
    ]
    return build_template_xlsx(headers, samples, "Template CBT")


def build_ubt_import_template_xlsx():
    headers = build_import_headers(UBT_REQUIRED_COLUMNS)
    samples = [
        ["UBT Bahasa Korea Dasar", "Latihan UBT dasar", "70", "1", "Apa arti sekolah?", "none", "", "Sekolah", "text", "", "Rumah", "text", "", "Pasar", "text", "", "Kantor", "text", "", "A", "hakgyo berarti sekolah."],
        ["UBT Bahasa Korea Dasar", "Latihan UBT dasar", "70", "2", "Pilih gambar yang sesuai.", "none", "", "", "image", "https://example.com/images/water.png", "Buku", "text", "", "Makanan", "text", "", "Teman", "text", "", "A", "Gambar air adalah jawaban yang benar."],
    ]
    return build_template_xlsx(headers, samples, "Template UBT")
